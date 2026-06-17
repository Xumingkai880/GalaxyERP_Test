"""
页面对象：导入订单页面 (Import Order Page)
功能：封装订单导入的核心 UI 操作
- 导航到手工订单页面
- 每次自动修改 Excel 订单号避免重复
- 点击"导入订单"按钮
- 上传 Excel 订单文件
- 确认导入并关闭结果弹窗
- 导入面单（导航待处理 → 批量操作 → 上传 PDF）
"""
import os
import random
import time
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from .base_page import BasePage


class ImportOrderPage(BasePage):
    """导入订单页面对象"""

    LOCATORS = {
        # 导航
        "order_menu": (By.ID, "tab-Order1912746753302405122"),
        "order_entry_menu": (
            By.XPATH,
            '//*[contains(@title, "Order entry") or contains(text(), "Order entry")]',
        ),
        "handmade_orders_menu": (
            By.XPATH,
            '//*[contains(@title, "Handmade orders") or contains(text(), "Handmade orders")]',
        ),
        # 导入订单按钮（多策略匹配）
        "add_order_btn": (
            By.XPATH,
            '//button[contains(., "导入订单") or contains(., "导入")]',
        ),
        # 对话框中的文件上传
        "import_file_input": (
            By.XPATH,
            '//div[@role="dialog"]//input[@type="file"]',
        ),
        # 确认按钮
        "confirm_btn": (
            By.XPATH,
            '//div[@role="dialog"]//button[contains(., "确定")]',
        ),
        # 待处理订单数量（验证用）
        "waiting_process_count": (
            By.XPATH,
            '//span[@title="Waiting Process"]/../span[@class="el-tag el-tag--primary el-tag--dark"]//span[@class="el-tag__content"]',
        ),
        # 面单导入相关
        "waiting_process_menu": (
            By.XPATH,
            '/html/body/div[1]/div/div[1]/div/div[1]/div/div[1]/div[1]/div/ul/li[1]/ul/li[1]',
        ),
        "bulk_operation_menu": (
            By.XPATH,
            '/html/body/div[1]/div/div[1]/div/div[1]/div/div[2]/div[2]/section/div/div[3]/div[1]/div/button',
        ),
    }

    # ==================== 导航 ====================

    def navigate_to_handmade_orders(self):
        """进入手工订单页面：订单标签 → 订单录入 → 手工订单"""
        self.find_and_click(self.LOCATORS["order_menu"])
        time.sleep(0.5)

        self.find_and_click(self.LOCATORS["order_entry_menu"])
        time.sleep(0.5)

        self.click_with_js(self.LOCATORS["handmade_orders_menu"])
        time.sleep(1)
        print("✅ 已进入手工订单页面")

    # ==================== 点击导入订单按钮 ====================

    def click_add_order(self):
        """点击'导入订单'按钮，打开导入对话框（多策略兜底）"""
        print("⏳ 尝试定位'导入订单'按钮...")

        # 先滚到顶部
        self.driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)

        # 策略 1：直接匹配"导入订单"文本
        locators = [
            (By.XPATH, '//button[contains(., "导入订单")]'),
            (By.XPATH, '//button[contains(., "导入")]'),
            (By.CSS_SELECTOR, '.el-button--primary'),
        ]

        for locator in locators:
            try:
                elements = self.driver.find_elements(*locator)
                for el in elements:
                    if el.is_displayed() and el.text.strip() in ("导入订单", "导入", ""):
                        self.driver.execute_script(
                            "arguments[0].click();", el
                        )
                        print("✅ 已点击导入订单按钮，对话框打开中...")
                        time.sleep(1)
                        return
            except Exception:
                continue

        raise Exception("❌ 未找到'导入订单'按钮，请确认页面加载完成")

    # ==================== 上传订单文件 ====================

    def upload_order_file(self, file_path: str):
        """上传 Excel 订单文件到对话框

        Args:
            file_path: Excel 订单文件的绝对路径
        """
        print(f"📂 准备上传订单文件: {file_path}")

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"订单文件不存在: {file_path}")

        time.sleep(0.5)

        # 定位对话框中的 file input
        file_inputs = self.driver.find_elements(
            By.XPATH, '//div[@role="dialog"]//input[@type="file"]'
        )
        if not file_inputs:
            file_inputs = self.driver.find_elements(
                By.CSS_SELECTOR, '.el-overlay-dialog input[type="file"]'
            )

        if not file_inputs:
            raise Exception("未找到对话框中的文件上传输入框")

        file_input = file_inputs[0]
        file_input.send_keys(file_path)
        time.sleep(2)

        file_name = file_input.get_attribute("value")
        if file_name:
            print(f"✅ 文件已选择: {file_name}")
        else:
            print("⚠️ 文件可能未成功选择")

    # ==================== 确认导入 ====================

    def confirm_import(self):
        """点击确定按钮，等待导入完成，关闭结果弹窗"""
        time.sleep(1)

        el = self.wait.until(
            EC.element_to_be_clickable(self.LOCATORS["confirm_btn"])
        )
        self.driver.execute_script("arguments[0].click();", el)
        print("✅ 已点击确定，等待导入完成...")
        time.sleep(3)

        self._close_result_dialog()

    def _close_result_dialog(self):
        """关闭导入结果弹窗（多策略兜底）"""
        try:
            print("⏳ 尝试关闭导入结果弹窗...")
            time.sleep(1)

            close_locators = [
                (By.XPATH, '//div[@class="el-dialog"]//footer//button[contains(., "关闭")]'),
                (By.XPATH, '//div[@class="dialog-footer"]//button'),
                (By.XPATH, '//button[contains(., "关闭")]'),
                (By.CSS_SELECTOR, '.el-dialog__footer button'),
            ]

            for locator in close_locators:
                try:
                    elements = self.driver.find_elements(*locator)
                    for el in elements:
                        if el.is_displayed():
                            self.driver.execute_script(
                                "arguments[0].scrollIntoView({block: 'center'});", el
                            )
                            time.sleep(0.2)
                            self.driver.execute_script("arguments[0].click();", el)
                            print("✅ 已关闭导入结果弹窗")
                            time.sleep(0.5)
                            return
                except Exception:
                    continue

            print("ℹ️ 未找到关闭按钮，可能已自动关闭")
        except Exception as e:
            print(f"⚠️ 关闭弹窗异常（不影响结果）: {e}")

    # ==================== 验证 ====================

    def get_order_count(self) -> int:
        """获取待处理订单数量"""
        try:
            time.sleep(0.5)
            elements = self.driver.find_elements(
                By.XPATH,
                '//span[@title="Waiting Process"]/../span[@class="el-tag el-tag--primary el-tag--dark"]//span[@class="el-tag__content"]',
            )
            if elements:
                count_text = elements[0].text.strip()
                if count_text and count_text.isdigit():
                    return int(count_text)
        except Exception:
            pass
        return 0

    # ==================== 订单号防重 ====================

    @staticmethod
    def _generate_unique_order_id() -> str:
        """生成唯一订单号：1163 + 15 位随机数字 + T，避免导入重复

        格式示例：1163876543210987654T  (19 位数字 + T)
        """
        prefix = "1163"
        digits = "".join(str(random.randint(0, 9)) for _ in range(15))
        return f"{prefix}{digits}T"

    @staticmethod
    def _update_excel_order_id(file_path: str, new_order_id: str):
        """修改 Excel 文件中的订单号（定位"订单号"列，更新第 2 行）

        Args:
            file_path: Excel 文件的绝对路径
            new_order_id: 新生成的唯一订单号
        """
        wb = load_workbook(file_path)
        ws = wb.active

        # 在表头行（第 1 行）中找"订单号"列
        target_col = 4  # 默认第 4 列
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "订单号":
                target_col = col
                break

        ws.cell(row=2, column=target_col, value=new_order_id)
        wb.save(file_path)
        print(f"  [OK] 已更新订单号为: {new_order_id}")

    # ==================== 完整流程入口 ====================

    def import_order(self, file_path: str):
        """执行完整的导入订单流程（每次自动修改订单号避免重复）

        Args:
            file_path: Excel 订单文件的绝对路径

        Steps:
            1. 导航到手工订单页面
            2. 记录导入前订单数
            3. 生成唯一订单号并更新 Excel
            4. 点击"导入订单"
            5. 上传文件
            6. 确认导入
            7. 打印导入后订单数
        """
        print("\n========== 开始导入订单 ==========")

        self.navigate_to_handmade_orders()

        before_count = self.get_order_count()
        print(f"  导入前待处理订单数: {before_count}")

        # 每次执行前随机修改订单号，避免重复
        new_id = self._generate_unique_order_id()
        self._update_excel_order_id(file_path, new_id)

        self.click_add_order()
        self.upload_order_file(file_path)
        self.confirm_import()

        after_count = self.get_order_count()
        print(f"  导入后待处理订单数: {after_count}")
        print(f"✅ 订单导入流程执行完毕（变化: {before_count} → {after_count}，订单号: {new_id}）\n")

    # ==================== 面单导入流程 ====================

    def navigate_to_waiting_process(self):
        """导航到待处理订单界面"""
        print("⏳ 导航到待处理订单界面...")
        el = self.wait.until(
            EC.element_to_be_clickable(self.LOCATORS["waiting_process_menu"])
        )
        self.driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", el
        )
        time.sleep(0.3)
        el.click()
        time.sleep(1)
        print("✅ 已进入待处理订单界面")

    def click_bulk_operation(self):
        """点击批量操作下拉框"""
        print("⏳ 点击批量操作下拉框...")
        el = self.wait.until(
            EC.element_to_be_clickable(self.LOCATORS["bulk_operation_menu"])
        )
        self.driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", el
        )
        time.sleep(0.3)
        el.click()
        # 等待下拉菜单展开
        try:
            self.wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, ".el-dropdown-menu__item")
                )
            )
            print("✅ 下拉菜单已展开")
        except Exception:
            print("⚠️ 下拉菜单未检测到，继续执行...")
        time.sleep(0.5)

    def click_import_waybill_menu(self):
        """在批量操作下拉框中点击'Import waybill'"""
        print("⏳ 点击 Import waybill 菜单项...")
        time.sleep(1)

        # 用 JS 遍历下拉菜单项找到并点击
        js_code = """
        var items = document.querySelectorAll('li.el-dropdown-menu__item');
        for (var i = 0; i < items.length; i++) {
            if (items[i].textContent.trim().includes('Import waybill')) {
                items[i].click();
                return true;
            }
        }
        return false;
        """
        result = self.driver.execute_script(js_code)
        if result:
            print("✅ 已点击 Import waybill 菜单项")
            time.sleep(1)
        else:
            # 兜底：用 Selenium 查找
            all_items = self.driver.find_elements(
                By.CSS_SELECTOR, "li.el-dropdown-menu__item"
            )
            print(f"🔍 找到 {len(all_items)} 个下拉菜单项")
            for item in all_items:
                if item.is_displayed() and "waybill" in item.text.lower():
                    self.driver.execute_script("arguments[0].click();", item)
                    time.sleep(1)
                    print("✅ 已点击导入运单菜单项")
                    return
            raise Exception("❌ 未找到'Import waybill'菜单项")

    def upload_waybill_file(self, file_path: str):
        """上传面单 PDF 文件到对话框

        Args:
            file_path: PDF 面单文件的绝对路径
        """
        print(f"📂 准备上传面单文件: {file_path}")

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"面单文件不存在: {file_path}")

        time.sleep(0.5)

        file_inputs = self.driver.find_elements(
            By.XPATH, '//div[@role="dialog"]//input[@type="file"]'
        )
        if not file_inputs:
            file_inputs = self.driver.find_elements(
                By.CSS_SELECTOR, '.el-overlay-dialog input[type="file"]'
            )

        if not file_inputs:
            raise Exception("未找到对话框中的面单文件上传输入框")

        file_input = file_inputs[0]
        file_input.send_keys(file_path)
        time.sleep(2)

        file_name = file_input.get_attribute("value")
        if file_name:
            print(f"✅ 面单文件已选择: {os.path.basename(file_name)}")
        else:
            print("⚠️ 面单文件可能未成功选择")

    def confirm_waybill_import(self):
        """确认面单导入（多策略定位 .el-button--primary 中的「确定」按钮）"""
        print("⏳ 定位面单确认按钮...")
        time.sleep(1)

        # 策略 1：按按钮类型 + 文本精确匹配
        try:
            btns = self.driver.find_elements(
                By.XPATH,
                '//button[contains(@class, "el-button--primary")]//span[text()="确定"]/parent::button',
            )
            for btn in btns:
                if btn.is_displayed():
                    self.driver.execute_script("arguments[0].click();", btn)
                    print("✅ 面单确认按钮已点击（策略1）")
                    time.sleep(3)
                    return
        except Exception:
            pass

        # 策略 2：dialog 范围内找 el-button--primary
        try:
            btns = self.driver.find_elements(
                By.XPATH,
                '//div[@role="dialog"]//button[contains(@class, "el-button--primary")]',
            )
            for btn in btns:
                if btn.is_displayed() and "确定" in btn.text:
                    self.driver.execute_script("arguments[0].click();", btn)
                    print("✅ 面单确认按钮已点击（策略2）")
                    time.sleep(3)
                    return
        except Exception:
            pass

        # 策略 3：全文兜底
        btns = self.driver.find_elements(
            By.XPATH, '//button[contains(@class, "el-button--primary")]'
        )
        for btn in btns:
            if btn.is_displayed() and ("确定" in btn.text or "确 定" in btn.text):
                self.driver.execute_script("arguments[0].click();", btn)
                print("✅ 面单确认按钮已点击（策略3）")
                time.sleep(3)
                return

        raise Exception("❌ 未找到面单导入的「确定」按钮")

    def import_waybill(self, file_path: str):
        """执行完整的面单导入流程

        Args:
            file_path: PDF 面单文件的绝对路径

        Steps:
            1. 导航到待处理订单界面
            2. 点击批量操作下拉框
            3. 点击 Import waybill 菜单项
            4. 上传面单文件
            5. 确认导入
        """
        print("\n========== 开始导入面单 ==========")

        self.navigate_to_waiting_process()
        self.click_bulk_operation()
        self.click_import_waybill_menu()
        self.upload_waybill_file(file_path)
        self.confirm_waybill_import()

        print("✅ 面单导入流程执行完毕\n")
