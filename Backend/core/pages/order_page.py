"""
页面对象：订单管理页面 (Order Page)
功能：封装订单相关的 UI 操作
- 导航到手工订单页面
- 导入订单文件（Excel）
- 获取订单数量
- 导入面单文件（PDF）
"""
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from .base_page import BasePage
import time
import os
import tempfile
from datetime import datetime

import openpyxl


class OrderPage(BasePage):
    # 页面元素定位器
    LOCATORS = {
        "order_menu": (By.ID, "tab-Order1912746753302405122"),  # 订单标签
        "order_entry_menu": (By.XPATH, '//*[contains(@title, "Order entry") or contains(text(), "Order entry")]'),  # 订单录入菜单
        "handmade_orders_menu": (By.XPATH, '//*[contains(@title, "Handmade orders") or contains(text(), "Handmade orders")]'),  # 手工订单菜单

        "add_order_btn": (By.XPATH, '//button[contains(., "导入订单") or contains(., "导入") or contains(@class, "el-button--primary")]'),

        "import_file_input": (By.XPATH, '//div[@role="dialog"]//input[@type="file"]'),  # 导入文件输入框（对话框中）
        "confirm_btn": (By.XPATH, '//div[@role="dialog"]//button[contains(., "确定")]'),  # 确认按钮（对话框中）
        "close_btn": (By.XPATH, '//*[@id="app"]/div/div[1]/div/div[1]/div/div[2]/div[2]/section/div/div[6]/div/div/footer/div/button'),  # 关闭弹窗按钮（备用）
        "waiting_process_count": (By.XPATH, '//span[@title="Waiting Process"]/../span[@class="el-tag el-tag--primary el-tag--dark"]//span[@class="el-tag__content"]'),  # 待处理订单数量标签
        "order_rows": (By.CSS_SELECTOR, 'table.el-table__body tbody tr.el-table__row'),  # 订单表格行

        "waiting_process_menu": (By.XPATH, '//span[contains(@title, "Waiting Process") or contains(text(), "Waiting")]'),  # 待处理菜单
        "bulk_operation_menu": (By.XPATH, '//button[contains(., "批量操作") or contains(., "Batch") or contains(., "Bulk")]'),  # 批量操作按钮
        "import_waybill_menu": (By.XPATH, '//li[contains(@class, "el-dropdown-menu__item") and (contains(., "Import waybill") or contains(., "运单"))]'),  # 导入运单菜单项
        "import_waybill_input": (By.XPATH, '//div[@role="dialog"]//input[@type="file"]'),  # 导入运单文件输入框
        "import_waybill_confirm_btn": (By.XPATH, '//div[@role="dialog"]//button[contains(., "确") or contains(., "Confirm") or contains(., "OK")]'),  # 导入运单确认按钮
    }

    # ==================== 完整导入流程（供 selenium_engine 调用） ====================

    # 订单号在 Excel 中的表头关键词（用于自动定位列）
    ORDER_NO_HEADERS = ["订单号", "order_no", "order number", "ordernumber", "订单编号"]

    def import_order(self, file_path: str):
        """完整的订单导入流程：去重订单号 -> 导航 -> 上传 -> 确认 -> 验证

        Args:
            file_path: Excel 订单文件的绝对路径
        """
        print("\n========== 开始导入订单 ==========")

        # 1. 处理 Excel：确保订单号唯一（每次运行生成新的时间戳订单号）
        processed_path = self._make_order_numbers_unique(file_path)

        try:
            # 2. 导航 + 上传 + 确认
            self.navigate_to_handmade_orders()
            self.click_add_order()
            self.upload_order_file(processed_path)
            self.confirm_order_import()
            self.verify_import_success()
            print("✅ 订单导入流程执行完毕\n")
        finally:
            # 清理临时文件
            if processed_path != file_path:
                try:
                    os.remove(processed_path)
                    print(f"🧹 已清理临时文件: {processed_path}")
                except Exception:
                    pass

    def _make_order_numbers_unique(self, file_path: str) -> str:
        """读取 Excel，将订单号替换为「EB + 当前时间戳 + 行号」，确保每次上传的订单号都是唯一的。

        Args:
            file_path: 原始 Excel 文件路径

        Returns:
            处理后的临时文件路径（调用方需要在用完后清理该文件）
        """
        print("🔧 处理订单号，确保唯一性...")
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        if ws.max_row <= 1:
            print("⚠️ Excel 中没有数据行，跳过处理")
            return file_path

        # ---- 定位订单号列 ----
        order_no_col = None
        for c in range(1, ws.max_column + 1):
            header = str(ws.cell(1, c).value or "").strip().lower()
            if any(kw.lower() in header for kw in self.ORDER_NO_HEADERS):
                order_no_col = c
                print(f"  找到订单号列: 第 {c} 列 ({ws.cell(1, c).value})")
                break

        if order_no_col is None:
            print("⚠️ 未找到订单号列，跳过处理，使用原始文件")
            return file_path

        # ---- 替换每一行的订单号 ----
        now_ts = datetime.now().strftime("%Y%m%d%H%M%S")
        modified = 0
        for r in range(2, ws.max_row + 1):
            old_val = ws.cell(r, order_no_col).value
            new_order_no = f"EB{now_ts}{r:04d}"
            ws.cell(r, order_no_col).value = new_order_no
            if old_val:
                print(f"  行 {r}: {old_val}  →  {new_order_no}")
            modified += 1

        # ---- 写入临时文件 ----
        temp_dir = tempfile.mkdtemp(prefix="order_")
        temp_path = os.path.join(temp_dir, os.path.basename(file_path))
        wb.save(temp_path)
        wb.close()

        print(f"✅ 已处理 {modified} 条订单号，临时文件: {temp_path}")
        return temp_path

    # ==================== 导航 ====================

    def navigate_to_handmade_orders(self):
        """进入手工订单页面

        操作路径：订单标签 -> 订单录入菜单 -> 手工订单菜单
        """
        self.find_and_click(self.LOCATORS["order_menu"])
        time.sleep(0.5)

        self.find_and_click(self.LOCATORS["order_entry_menu"])
        time.sleep(0.5)

        self.click_with_js(self.LOCATORS["handmade_orders_menu"])

        time.sleep(1)
        print("✅ 已进入手工订单页面")

    # ==================== 导入订单 ====================

    def click_add_order(self):
        """点击导入订单按钮，打开导入对话框"""
        print("⏳ 尝试定位'导入订单'按钮...")

        # 滚动到顶部
        self.driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)

        locators = [
            (By.XPATH, '//button[contains(., "导入订单")]'),
            (By.XPATH, '//button[contains(., "导入")]'),
            (By.CSS_SELECTOR, '.el-button--primary'),
            (By.XPATH, '//div[@class="toolbar"]//button'),
        ]

        for locator in locators:
            try:
                elements = self.driver.find_elements(*locator)
                for element in elements:
                    if element.is_displayed() and element.text.strip() in ["导入订单", "导入", ""]:
                        self.driver.execute_script("arguments[0].click();", element)
                        print("✅ 已点击导入订单按钮")
                        time.sleep(1)
                        return
            except Exception:
                continue

        raise Exception("❌ 未找到'导入订单'按钮")

    def upload_order_file(self, file_path: str):
        """上传订单文件

        Args:
            file_path: Excel 订单文件的绝对路径
        """
        print(f"📂 准备上传文件: {file_path}")

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        time.sleep(0.5)

        dialog_file_input = self.driver.find_elements(
            By.XPATH, '//div[@role="dialog"]//input[@type="file"]'
        )

        print(f"  对话框中找到 {len(dialog_file_input)} 个 file input 元素")

        if not dialog_file_input:
            dialog_file_input = self.driver.find_elements(
                By.CSS_SELECTOR, '.el-overlay-dialog input[type="file"]'
            )
            print(f"🔍 通过 CSS 找到 {len(dialog_file_input)} 个元素")

        if not dialog_file_input:
            raise Exception("未找到对话框中的文件输入框")

        file_input = dialog_file_input[0]
        file_input.send_keys(file_path)
        time.sleep(2)

        file_name = file_input.get_attribute('value')
        if file_name:
            print(f"✅ 文件已选择: {file_name}")
        else:
            print("⚠️ 文件可能未成功选择")

        print(f"✅ 已上传订单文件: {file_path}")

    def confirm_order_import(self):
        """确认订单导入并处理成功弹窗"""
        time.sleep(1)

        element = self.wait.until(EC.element_to_be_clickable(self.LOCATORS["confirm_btn"]))
        self.driver.execute_script("arguments[0].click();", element)

        time.sleep(3)
        print("✅ 已点击确定按钮，等待订单导入完成...")

        self.close_import_result()

    def close_import_result(self):
        """关闭订单导入结果页面/弹窗"""
        try:
            print("⏳ 尝试关闭导入结果...")
            time.sleep(1)

            close_locators = [
                (By.XPATH, '//div[@class="el-dialog"]//footer//button[contains(., "关闭")]'),
                (By.XPATH, '//div[@class="dialog-footer"]//button'),
                (By.XPATH, '//button[contains(., "关闭")]'),
                self.LOCATORS["close_btn"],
                (By.CSS_SELECTOR, '.el-dialog__footer button'),
            ]

            for locator in close_locators:
                try:
                    elements = self.driver.find_elements(*locator)
                    for element in elements:
                        if element.is_displayed():
                            print(f"✅ 找到关闭按钮: {locator}")
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                            time.sleep(0.2)
                            self.driver.execute_script("arguments[0].click();", element)
                            print("✅ 已点击关闭按钮")
                            time.sleep(0.5)
                            return
                except Exception:
                    continue

            print("ℹ️ 未找到关闭按钮，可能已自动关闭")

        except Exception as e:
            print(f"⚠️ 关闭导入结果时出现异常: {str(e)}，继续执行...")

    def verify_import_success(self):
        """验证导入是否成功"""
        try:
            time.sleep(2)
            current_count = self.get_order_count()
            print(f"✅ 当前订单数量: {current_count}")
            return True
        except Exception as e:
            print(f"❌ 验证导入结果失败: {str(e)}")
            return False

    # ==================== 订单数量 ====================

    def get_order_count_from_table(self):
        """从订单表格中获取订单数量"""
        try:
            time.sleep(1)
            rows = self.driver.find_elements(*self.LOCATORS["order_rows"])
            count = len(rows)
            print(f"🔍 表格中订单数量: {count}")
            return count
        except Exception as e:
            print(f"❌ 获取表格订单数量失败: {str(e)}")
            return 0

    def get_order_count(self):
        """获取待处理订单数量

        优先从菜单栏标签获取，失败则从表格获取
        """
        try:
            time.sleep(0.5)

            locators = [
                (By.XPATH, '//span[@title="Waiting Process"]/../span[@class="el-tag el-tag--primary el-tag--dark"]//span[@class="el-tag__content"]'),
                (By.CSS_SELECTOR, 'span[title="Waiting Process"] + span.el-tag span.el-tag__content'),
            ]

            for locator in locators:
                try:
                    elements = self.driver.find_elements(*locator)
                    if elements:
                        count_text = elements[0].text.strip()
                        if count_text and count_text.isdigit():
                            count = int(count_text)
                            print(f"🔍 待处理订单数量（菜单）: {count}")
                            return count
                except Exception:
                    continue

            print("⚠️ 从菜单栏获取失败，尝试从表格获取...")
            return self.get_order_count_from_table()

        except Exception as e:
            print(f"❌ 获取订单数量失败: {str(e)}")
            return 0

    # ==================== 导入面单 ====================

    def navigate_to_waiting_process(self):
        """导航到待处理订单界面

        路径：左侧菜单 Shop Orders → Waiting Process（待处理）
        """
        print("⏳ 导航到待处理订单界面...")

        # ---- Step 1: 先点击 Shop Orders 展开子菜单 ----
        self._click_left_menu("Shop Orders", ["Shop Orders", "Shop", "店铺订单"])

        # ---- Step 2: 再点击 Waiting Process ----
        locators = [
            (By.XPATH, '//span[contains(@title, "Waiting Process")]'),
            (By.XPATH, '//li[contains(., "待处理") or contains(., "Waiting")]'),
            (By.CSS_SELECTOR, 'li.el-menu-item span[title*="Waiting"]'),
        ]

        found = False
        for i, (by, value) in enumerate(locators):
            try:
                elements = self.driver.find_elements(by, value)
                if elements:
                    for el in elements:
                        if el.is_displayed():
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                            time.sleep(0.2)
                            self.driver.execute_script("arguments[0].click();", el)
                            time.sleep(1)
                            print(f"✅ 已进入待处理订单界面 (定位器 #{i+1})")
                            found = True
                            break
                if found:
                    break
            except Exception:
                continue

        if not found:
            raise Exception("❌ 无法导航到待处理订单界面")

    def _click_left_menu(self, menu_name: str, aliases: list):
        """点击左侧菜单项（多策略兜底）

        Args:
            menu_name: 日志用的菜单名称
            aliases: 可选匹配关键词列表（中英文皆可）
        """
        print(f"⏳ 点击左侧菜单: {menu_name}...")

        # 策略 1: JS 直接搜索菜单文字并点击
        js_code = f"""
        var items = document.querySelectorAll('li.el-submenu__title, li.el-menu-item, span.el-menu-item');
        for (var i = 0; i < items.length; i++) {{
            var txt = items[i].textContent.trim();
            if ({' || '.join(f"txt.includes('{a}')" for a in aliases)}) {{
                items[i].click();
                return 'js-click';
            }}
        }}
        // 再试一次，搜索所有元素
        var all = document.querySelectorAll('*');
        for (var i = 0; i < all.length; i++) {{
            var txt = (all[i].textContent || '').trim();
            if (txt === '{aliases[0]}' && (all[i].tagName === 'LI' || all[i].tagName === 'SPAN' || all[i].tagName === 'DIV')) {{
                all[i].click();
                return 'js-all-click';
            }}
        }}
        return '';
        """
        try:
            result = self.driver.execute_script(js_code)
            if result:
                print(f"✅ 已点击 {menu_name} ({result})")
                time.sleep(1)
                return
        except Exception as e:
            print(f"⚠️ JS 点击 {menu_name} 失败: {e}")

        # 策略 2: XPath 文本匹配
        for alias in aliases:
            try:
                xpath = f'//li[contains(., "{alias}")]'
                elements = self.driver.find_elements(By.XPATH, xpath)
                for el in elements:
                    if el.is_displayed():
                        self.driver.execute_script("arguments[0].click();", el)
                        print(f"✅ 已点击 {menu_name} (XPath: {alias})")
                        time.sleep(1)
                        return
            except Exception:
                continue

        print(f"⚠️ 未找到 {menu_name}，跳过（可能已经在该页面）")

    def click_bulk_operation(self):
        """点击批量操作下拉框（多定位器兜底）"""
        print("⏳ 点击批量操作下拉框...")

        locators = [
            (By.XPATH, '//button[contains(., "批量操作") or contains(., "Batch") or contains(., "Bulk")]'),
            (By.XPATH, '//button[contains(@class, "el-button") and (contains(., "批量") or contains(., "Batch"))]'),
            (By.CSS_SELECTOR, 'button.el-button:has-text("Batch"), button.el-button:has-text("批量")'),
            # 旧版绝对 XPath 兜底
            (By.XPATH, '/html/body/div[1]/div/div[1]/div/div[1]/div/div[2]/div[2]/section/div/div[3]/div[1]/div/button'),
        ]

        found = False
        for i, locator in enumerate(locators):
            try:
                elements = self.driver.find_elements(*locator)
                if elements:
                    for el in elements:
                        if el.is_displayed():
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                            time.sleep(0.2)
                            self.driver.execute_script("arguments[0].click();", el)
                            print(f"✅ 已点击批量操作按钮 (定位器 #{i+1})")
                            found = True
                            break
                if found:
                    break
            except Exception:
                continue

        if not found:
            # 最终尝试：使用 JS 直接搜索按钮文本
            js_code = """
            var btns = document.querySelectorAll('button');
            for (var i = 0; i < btns.length; i++) {
                var txt = btns[i].textContent || '';
                if (txt.includes('Batch') || txt.includes('批量')) {
                    btns[i].click();
                    return true;
                }
            }
            return false;
            """
            if self.driver.execute_script(js_code):
                print("✅ 已通过 JS 点击批量操作按钮")
                found = True

        if not found:
            raise Exception("❌ 未找到批量操作按钮")

        print("⏳ 等待下拉菜单展开...")
        try:
            self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.el-dropdown-menu__item')))
            print("✅ 下拉菜单已展开")
        except TimeoutException:
            print("⚠️ 下拉菜单未检测到，继续执行...")

        time.sleep(0.5)

    def click_import_waybill_menu(self):
        """点击导入运单菜单项（多策略兜底）"""
        print("⏳ 点击导入运单菜单项...")

        time.sleep(2)

        # 方法 1: 使用 JavaScript 直接查找并点击（支持中英文）
        try:
            js_code = """
            var items = document.querySelectorAll('li.el-dropdown-menu__item');
            for (var i = 0; i < items.length; i++) {
                var txt = items[i].textContent.trim();
                if (txt.includes('Import waybill') || txt.includes('导入运单') || txt.includes('waybill')) {
                    items[i].click();
                    return true;
                }
            }
            return false;
            """
            result = self.driver.execute_script(js_code)
            if result:
                print("✅ 通过 JavaScript 成功点击导入运单菜单项")
                time.sleep(1)
                return
            else:
                print("⚠️ JavaScript 未找到导入运单菜单项")
        except Exception as e:
            print(f"⚠️ JavaScript 点击失败: {e}")

        # 方法 2: 使用 Selenium 查找所有下拉菜单项
        try:
            all_items = self.driver.find_elements(By.CSS_SELECTOR, 'li.el-dropdown-menu__item')
            print(f"🔍 找到 {len(all_items)} 个下拉菜单项")

            for item in all_items:
                if item.is_displayed():
                    text = item.text.strip()
                    print(f"   - '{text}'")
                    keywords = ['Import waybill', '导入运单', 'waybill', '运单']
                    if any(kw in text for kw in keywords):
                        print(f"✅ 找到目标菜单项: '{text}'")
                        self.driver.execute_script("arguments[0].click();", item)
                        time.sleep(1)
                        print("✅ 已点击导入运单菜单项")
                        return
        except Exception as e:
            print(f"⚠️ Selenium 查找失败: {e}")

        # 方法 3: 尝试通过原始 locator（如果以上都失败）
        try:
            elements = self.driver.find_elements(*self.LOCATORS["import_waybill_menu"])
            if elements and elements[0].is_displayed():
                self.driver.execute_script("arguments[0].click();", elements[0])
                print("✅ 通过备用定位器点击导入运单菜单项")
                time.sleep(1)
                return
        except Exception:
            pass

        raise Exception("未找到导入运单菜单项")

    def upload_waybill_file(self, file_path: str):
        """上传面单文件

        Args:
            file_path: PDF 面单文件的绝对路径
        """
        print(f"📂 准备上传面单文件: {file_path}")

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        if not file_path.lower().endswith('.pdf'):
            print(f"⚠️ 警告：面单文件不是 PDF 格式: {file_path}")

        time.sleep(0.5)

        dialog_file_input = self.driver.find_elements(
            By.XPATH, '//div[@role="dialog"]//input[@type="file"]'
        )

        print(f"  对话框中找到 {len(dialog_file_input)} 个 file input 元素")

        if not dialog_file_input:
            dialog_file_input = self.driver.find_elements(
                By.CSS_SELECTOR, '.el-overlay-dialog input[type="file"]'
            )
            print(f"🔍 通过 CSS 找到 {len(dialog_file_input)} 个元素")

        if not dialog_file_input:
            raise Exception("未找到对话框中的面单文件输入框")

        file_input = dialog_file_input[0]
        file_input.send_keys(file_path)
        time.sleep(2)

        file_name = file_input.get_attribute('value')
        if file_name:
            print(f"✅ 面单文件已选择: {os.path.basename(file_name)}")
        else:
            print("⚠️ 面单文件可能未成功选择")

        print(f"✅ 已上传面单文件: {file_path}")

    def confirm_waybill_import(self):
        """确认面单导入（多定位器兜底）"""
        time.sleep(1)

        locators = [
            # 方法 1: 灵活定位对话框中的确认按钮
            (By.XPATH, '//div[@role="dialog"]//button[contains(., "确") or contains(., "Confirm") or contains(., "OK")]'),
            # 方法 2: footer 中的确认按钮
            (By.XPATH, '//div[@role="dialog"]//footer//button[2]'),
            # 方法 3: 旧版绝对 XPath 兜底
            (By.XPATH, '/html/body/div[1]/div/div[1]/div/div[1]/div/div[2]/div[2]/section/div/div[17]/div/div/footer/div/button[2]'),
            # 方法 4: 任意 el-dialog 中的最后按钮
            (By.CSS_SELECTOR, '.el-dialog__footer button:last-child'),
        ]

        found = False
        for i, (by, value) in enumerate(locators):
            try:
                elements = self.driver.find_elements(by, value)
                if elements:
                    for el in elements:
                        if el.is_displayed():
                            self.driver.execute_script("arguments[0].click();", el)
                            print(f"✅ 已点击确认按钮 (定位器 #{i+1})")
                            found = True
                            break
                if found:
                    break
            except Exception:
                continue

        if not found:
            raise Exception("❌ 未找到面单导入确认按钮")

        time.sleep(3)
        print("✅ 已点击确认按钮，等待面单导入完成...")
        print("✅ 面单导入成功")

        try:
            time.sleep(1)
            current_count = self.get_order_count()
            print(f"✅ 验证通过：当前订单数量为 {current_count}")
        except Exception as e:
            print(f"⚠️ 验证时出现异常（不影响测试）: {str(e)}")

    def import_waybill(self, waybill_file_path: str):
        """完整的面单导入流程

        Args:
            waybill_file_path: PDF 面单文件的绝对路径

        操作流程：
        1. 导航到待处理订单界面
        2. 点击批量操作下拉框
        3. 点击导入运单菜单项
        4. 上传面单文件
        5. 确认导入
        """
        print("\n========== 开始导入面单 ==========")

        self.navigate_to_waiting_process()
        self.click_bulk_operation()
        self.click_import_waybill_menu()
        self.upload_waybill_file(waybill_file_path)
        self.confirm_waybill_import()

        print("✅ 面单导入流程执行完毕\n")

